"""Utilities for securely validating and parsing customer import uploads."""

import io
import os
import zipfile
from dataclasses import dataclass
from typing import Tuple

import pandas as pd
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB hard cap to prevent resource exhaustion.
MAX_ROWS = 100_000  # Prevent excessively large imports that could choke the database.
CHUNK_SIZE = 5_000  # Stream CSV parsing to avoid loading huge files in memory.
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}


@dataclass
class ParsedUpload:
    dataframe: pd.DataFrame
    filename: str
    file_size: int
    row_count: int


class UploadError(Exception):
    """Raised when an uploaded file fails validation."""

    def __init__(self, user_message: str, status_code: int, *, log_message: str = None, filename: str = "") -> None:
        super().__init__(user_message)
        self.user_message = user_message
        self.status_code = status_code
        self.log_message = log_message or user_message
        self.filename = filename


class UploadTooLargeError(UploadError):
    """Specific error for oversized files."""

    def __init__(self, user_message: str, *, filename: str = "") -> None:
        super().__init__(user_message, 413, log_message=user_message, filename=filename)


def parse_import_file(file_storage: FileStorage) -> ParsedUpload:
    """Validate and parse an uploaded CSV or XLSX file into a DataFrame."""

    original_filename = file_storage.filename or ""
    filename = secure_filename(original_filename)
    if not filename:
        raise UploadError("No file selected for uploading.", 400, filename=filename)

    _, ext = os.path.splitext(filename)
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise UploadError("Invalid file type. Please upload a .csv or .xlsx file.", 415, filename=filename)

    file_storage.stream.seek(0)
    data = file_storage.stream.read(MAX_UPLOAD_SIZE + 1)
    if not data:
        raise UploadError("Uploaded file is empty.", 400, filename=filename)

    if len(data) > MAX_UPLOAD_SIZE:
        raise UploadTooLargeError("File exceeds the 10 MB upload limit.", filename=filename)

    if ext == ".csv":
        dataframe, row_count = _parse_csv_bytes(data)
    else:
        dataframe, row_count = _parse_xlsx_bytes(data)

    return ParsedUpload(
        dataframe=dataframe,
        filename=filename,
        file_size=len(data),
        row_count=row_count,
    )


def _parse_csv_bytes(data: bytes) -> Tuple[pd.DataFrame, int]:
    snippet = data[:4096]
    if b"\x00" in snippet:
        raise UploadError("Uploaded CSV appears to contain binary data and was rejected.", 415)

    encoding = "utf-8"
    try:
        snippet_text = snippet.decode("utf-8")
    except UnicodeDecodeError:
        snippet_text = snippet.decode("latin-1")
        encoding = "latin-1"

    if not any(sep in snippet_text for sep in (",", ";", "\t")):
        raise UploadError("Uploaded CSV does not contain a recognised delimiter (comma, semicolon, or tab).", 400)

    buffer = io.BytesIO(data)
    text_wrapper = io.TextIOWrapper(buffer, encoding=encoding, newline="")
    frames = []
    total_rows = 0
    try:
        for chunk in pd.read_csv(text_wrapper, dtype=str, chunksize=CHUNK_SIZE):
            total_rows += len(chunk)
            if total_rows > MAX_ROWS:
                raise UploadTooLargeError("CSV contains more than 100000 rows and was rejected.")
            frames.append(chunk)
    except UploadError:
        raise
    except Exception as exc:
        raise UploadError(f"Failed to parse CSV: {exc}", 400) from exc
    finally:
        text_wrapper.detach()

    dataframe = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    dataframe = _strip_formula_injection(dataframe)
    return dataframe, total_rows


def _parse_xlsx_bytes(data: bytes) -> Tuple[pd.DataFrame, int]:
    buffer = io.BytesIO(data)
    if not zipfile.is_zipfile(buffer):
        raise UploadError("Uploaded XLSX is not a valid Excel file.", 415)

    with zipfile.ZipFile(buffer, "r") as archive:
        names = archive.namelist()
        lowered = {name.lower() for name in names}
        if "[content_types].xml" not in lowered:
            raise UploadError("Uploaded XLSX is missing required metadata and was rejected.", 415)
        if any(name.lower().endswith("vbaproject.bin") for name in names):
            raise UploadError("Excel macros are not allowed in uploaded files.", 415)

    buffer.seek(0)
    try:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
    except Exception as exc:
        raise UploadError(f"Failed to open XLSX file: {exc}", 400) from exc

    rows = []
    row_count = 0
    try:
        sheet = workbook.active
        headers = None
        for row in sheet.iter_rows(values_only=True):
            if headers is None:
                headers = [_normalise_header(cell) for cell in row]
                continue

            if all(cell is None for cell in row):
                continue

            row_count += 1
            if row_count > MAX_ROWS:
                raise UploadTooLargeError("XLSX contains more than 100000 rows and was rejected.")

            record = {}
            for header, cell in zip(headers, row):
                if not header:
                    continue
                value = _strip_leading_equals(cell) if isinstance(cell, str) else cell
                record[header] = value
            rows.append(record)
    except UploadError:
        raise
    except Exception as exc:
        raise UploadError(f"Failed to parse XLSX: {exc}", 400) from exc
    finally:
        workbook.close()

    dataframe = pd.DataFrame(rows)
    dataframe = _strip_formula_injection(dataframe)
    return dataframe, row_count


def _normalise_header(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, str):
        return cell.strip()
    return str(cell).strip()


def _strip_formula_injection(dataframe: pd.DataFrame) -> pd.DataFrame:
    if dataframe.empty:
        return dataframe

    for column in dataframe.select_dtypes(include=["object"]):
        dataframe[column] = dataframe[column].apply(_strip_leading_equals)
    return dataframe


def _strip_leading_equals(value):
    if isinstance(value, str) and value.startswith("="):
        return value.lstrip("=")
    return value